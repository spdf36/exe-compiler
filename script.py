import datetime
import os
from io import BytesIO

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image


def get_extension(filename):
    _, ext = os.path.splitext(filename.lower())
    return ext


def analyze_local_images(target_folder):
    """
    Recursively scans a local folder and runs image extension checks.
    Returns a list of structured row data for our spreadsheet.
    """
    extension_vs_format = {".jpg": "jpeg", ".jpeg": "jpeg", ".png": "png"}

    supported_extensions = list(extension_vs_format.keys())
    results = []

    print(f"\nScanning target directory: {target_folder}...")

    # os.walk scans recursively
    for root, dirs, files in os.walk(target_folder):
        for file in files:
            ext = get_extension(file)
            if ext not in supported_extensions:
                continue  # Skip non-image files

            full_path = os.path.join(root, file)
            # Use relative path for cleaner reading inside the Excel report
            rel_path = os.path.relpath(full_path, target_folder)

            try:
                with open(full_path, "rb") as f:
                    image_bytes = f.read()

                image = Image.open(BytesIO(image_bytes))
                actual_format = image.format.lower()
                expected_format = extension_vs_format.get(ext)

                if expected_format != actual_format:
                    status = "MISMATCH"
                    details = f"File has extension '{ext}' but actual format is '{actual_format}'"
                else:
                    status = "PASS"
                    details = "Extension matches binary file structure perfectly."

            except Exception as e:
                actual_format = "unknown"
                # HEIC Magic Byte Check
                header = image_bytes[:20] if "image_bytes" in locals() else b""
                if b"ftypheic" in header or b"ftypheif" in header:
                    status = "HEIC/HEIF FORMAT"
                    details = f"File has extension '{ext}' but actual binary format is HEIC/HEIF (iPhone Raw)"
                else:
                    status = "UNREADABLE"
                    details = f"File could not be validated: {str(e)}"

            results.append((file, rel_path, ext, actual_format, status, details))

    return results


def generate_styled_excel(data, output_filepath):
    """
    Generates a beautifully styled corporate Excel workbook with automated KPIs,
    color-coded conditional statuses, and a tracking chart.
    """
    wb = openpyxl.Workbook()

    # 1. Setup worksheets
    ws_summary = wb.active
    ws_summary.title = "Summary Dashboard"
    ws_details = wb.create_sheet(title="Scan Results")

    # Ensure grid lines are visible
    ws_summary.views.sheetView[0].showGridLines = True
    ws_details.views.sheetView[0].showGridLines = True

    # Define Corporate Style Palettes (Muted Navy, Soft Alerts)
    NAVY_HEADER = "1F4E79"
    WHITE = "FFFFFF"
    ZEBRA_GRAY = "F2F4F4"
    BORDER_GRAY = "D5D8DC"

    PASS_FILL, PASS_TEXT = "E8F8F5", "117A65"  # Soft Pastel Green
    FAIL_FILL, FAIL_TEXT = "FADBD8", "78281F"  # Soft Pastel Red
    WARN_FILL, WARN_TEXT = "FEF9E7", "7D6608"  # Soft Pastel Yellow

    font_title = Font(name="Segoe UI", size=16, bold=True, color=NAVY_HEADER)
    font_header = Font(name="Segoe UI", size=11, bold=True, color=WHITE)
    font_body = Font(name="Segoe UI", size=10)
    font_body_bold = Font(name="Segoe UI", size=10, bold=True)

    fill_header = PatternFill(
        start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid"
    )
    fill_zebra = PatternFill(
        start_color=ZEBRA_GRAY, end_color=ZEBRA_GRAY, fill_type="solid"
    )

    thin_border = Border(
        left=Side(style="thin", color=BORDER_GRAY),
        right=Side(style="thin", color=BORDER_GRAY),
        top=Side(style="thin", color=BORDER_GRAY),
        bottom=Side(style="thin", color=BORDER_GRAY),
    )

    # --- Populating Detailed Sheet ---
    headers = [
        "File Name",
        "Relative Path",
        "Extension",
        "Detected Format",
        "Status",
        "Details",
    ]
    ws_details.append(headers)

    for row in data:
        ws_details.append(row)

    # Apply styling & layouts to the results sheet
    for col_num in range(1, len(headers) + 1):
        cell = ws_details.cell(row=1, column=col_num)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    total_rows = len(data) + 1
    for row_idx in range(2, total_rows + 1):
        status = ws_details.cell(row=row_idx, column=5).value
        is_zebra = row_idx % 2 == 0

        for col_idx in range(1, len(headers) + 1):
            cell = ws_details.cell(row=row_idx, column=col_idx)
            cell.font = font_body
            cell.border = thin_border
            if is_zebra:
                cell.fill = fill_zebra

            # Content-driven cell formats
            if col_idx == 5:  # Status Column
                cell.alignment = Alignment(horizontal="center")
                if status == "PASS":
                    cell.fill = PatternFill(
                        start_color=PASS_FILL, end_color=PASS_FILL, fill_type="solid"
                    )
                    cell.font = Font(
                        name="Segoe UI", size=10, bold=True, color=PASS_TEXT
                    )
                elif status == "MISMATCH":
                    cell.fill = PatternFill(
                        start_color=FAIL_FILL, end_color=FAIL_FILL, fill_type="solid"
                    )
                    cell.font = Font(
                        name="Segoe UI", size=10, bold=True, color=FAIL_TEXT
                    )
                else:
                    cell.fill = PatternFill(
                        start_color=WARN_FILL, end_color=WARN_FILL, fill_type="solid"
                    )
                    cell.font = Font(
                        name="Segoe UI", size=10, bold=True, color=WARN_TEXT
                    )
            elif col_idx in [3, 4]:
                cell.alignment = Alignment(horizontal="center")

    # Set smart column widths dynamically
    for col in ws_details.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_details.column_dimensions[col_letter].width = max(max_len + 3, 12)

    ws_details.freeze_panes = "A2"
    ws_details.auto_filter.ref = f"A1:F{total_rows}"

    # --- Populating Summary Dashboard Sheet ---
    ws_summary.cell(
        row=2, column=2, value="Image Format Validation Dashboard"
    ).font = font_title
    ws_summary.cell(
        row=3,
        column=2,
        value=f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
    ).font = Font(name="Segoe UI", size=10, italic=True, color="555555")

    ws_summary.cell(row=5, column=2, value="Validation Metric").font = font_header
    ws_summary.cell(row=5, column=2).fill = fill_header
    ws_summary.cell(row=5, column=3, value="Count").font = font_header
    ws_summary.cell(row=5, column=3).fill = fill_header
    ws_summary.cell(row=5, column=3).alignment = Alignment(horizontal="right")

    metrics = [
        ("Total Images Scanned", f"=COUNTA('Scan Results'!A2:A{total_rows})"),
        ("Passed (Valid)", f"=COUNTIF('Scan Results'!E2:E{total_rows}, \"PASS\")"),
        (
            "Mismatched Extensions",
            f"=COUNTIF('Scan Results'!E2:E{total_rows}, \"MISMATCH\")",
        ),
        (
            "HEIC/HEIF Warnings",
            f"=COUNTIF('Scan Results'!E2:E{total_rows}, \"HEIC/HEIF FORMAT\")",
        ),
        (
            "Unreadable/Corrupt",
            f"=COUNTIF('Scan Results'!E2:E{total_rows}, \"UNREADABLE\")",
        ),
    ]

    for idx, (metric, formula) in enumerate(metrics, start=6):
        c1 = ws_summary.cell(row=idx, column=2, value=metric)
        c2 = ws_summary.cell(row=idx, column=3, value=formula)
        c1.font = font_body if idx != 6 else font_body_bold
        c2.font = font_body if idx != 6 else font_body_bold
        c1.border = thin_border
        c2.border = thin_border
        c2.alignment = Alignment(horizontal="right")
        if idx == 6:
            compress_fill = fill_zebra
            c1.fill = compress_fill
            c2.fill = compress_fill

    ws_summary.column_dimensions["B"].width = 28
    ws_summary.column_dimensions["C"].width = 12

    # 3. Add dynamic Chart component to Dashboard
    if len(data) > 0:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Validation Status Breakdown"
        chart.y_axis.title = "Number of Files"
        chart.x_axis.title = "Status Group"

        data_ref = Reference(ws_summary, min_col=3, min_row=6, max_row=10)
        cats_ref = Reference(
            ws_summary, min_col=2, min_row=7, max_row=10
        )  # Drop down to status splits

        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(cats_ref)
        chart.legend = None
        chart.width = 16
        chart.height = 10
        ws_summary.add_chart(chart, "E5")

    wb.save(output_filepath)
    print(f"✨ Success! Report saved cleanly to: {output_filepath}")


# --- Interactive Execution Entrypoint ---
if __name__ == "__main__":
    print("====================================================")
    print("        IMAGE FORMAT SPECIFICATION VERIFIER        ")
    print("====================================================\n")

    # 1. Prompt for target directory with validation loop
    while True:
        target_dir = input("📂 Drag & drop or enter the folder path to scan: ").strip()
        
        # Clean up quotation marks automatically added by terminal drag-and-drop
        target_dir = target_dir.strip("'\"")
        
        if os.path.isdir(target_dir):
            break
        print("❌ That path doesn't seem to exist. Please try again.\n")

    # 2. Prompt for output report name
    report_name = input("📝 Enter output spreadsheet name (Default: Image_Validation_Report): ").strip()
    if not report_name:
        report_name = "Image_Validation_Report"
        
    # Append extension automatically if missing
    if not report_name.lower().endswith(".xlsx"):
        report_name += ".xlsx"

    print("\n----------------------------------------------------")

    # 3. Run Pipeline execution
    try:
        scan_data = analyze_local_images(target_dir)
        if scan_data:
            generate_styled_excel(scan_data, report_name)
        else:
            print("\n⚠️ No supported image profiles (.png, .jpg, .jpeg) found in target folder.")
    except Exception as e:
        print(f"\n❌ A critical error halted the processing loop: {e}")

    # 4. Prevent the terminal window from automatically vanishing
    print("\n====================================================")
    input("🏁 Process finished. Press [ENTER] to exit close this window...")
